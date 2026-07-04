"""
OCR·텍스트 추출 모듈
PDF, 이미지, Excel, HWP 파일에서 로컬 텍스트 추출
"""

import logging
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

import pdfplumber
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
from pdf2image import convert_from_path

import config

logger = logging.getLogger(__name__)

# Tesseract OCR 언어
OCR_LANG: str = "kor+eng"

# PDF 페이지 분할 기준
PDF_PAGE_CHUNK_SIZE: int = 50
PDF_PAGE_OCR_LIMIT: int = 100

# 품질 경고 임계값
QUALITY_WARN_THRESHOLD: float = 0.3

# LibreOffice 실행 파일 후보 경로 (Windows, config.LIBREOFFICE_PATH 우선)
LIBREOFFICE_PATHS: list[str] = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]

LIBREOFFICE_HWP_FILTERS: list[str | None] = [
    "Hwp2002_File",
    "Hwp2002_Reader",
    None,
]


def _empty_result(error: str) -> dict[str, Any]:
    """실패 결과 템플릿"""
    return {
        "success": False,
        "method": None,
        "text": "",
        "page_count": 0,
        "char_count": 0,
        "ocr_used": False,
        "quality_score": 0.0,
        "error": error,
    }


def _success_result(
    text: str,
    method: str,
    page_count: int = 0,
    ocr_used: bool = False,
) -> dict[str, Any]:
    """성공 결과 템플릿"""
    quality = check_ocr_quality(text)
    if quality < QUALITY_WARN_THRESHOLD:
        logger.warning(
            "텍스트 품질 점수 낮음: %.2f (임계값 %.2f)",
            quality,
            QUALITY_WARN_THRESHOLD,
        )
    return {
        "success": True,
        "method": method,
        "text": text,
        "page_count": page_count,
        "char_count": len(text),
        "ocr_used": ocr_used,
        "quality_score": quality,
        "error": None,
    }


def check_ocr_quality(text: str) -> float:
    """
    추출 텍스트 품질 점수 반환 (0~1).
    한글 비율, 텍스트 길이, 특수문자 비율 등을 종합 평가.
    """
    if not text or not text.strip():
        return 0.0

    cleaned = text.strip()
    total = len(cleaned)
    if total == 0:
        return 0.0

    # 한글 문자 비율
    korean_chars = len(re.findall(r"[가-힣]", cleaned))
    korean_ratio = korean_chars / total

    # 알파벳·숫자 비율
    alnum_chars = len(re.findall(r"[A-Za-z0-9]", cleaned))
    alnum_ratio = alnum_chars / total

    # 의미 있는 문자 비율 (한글+영문+숫자+공백)
    meaningful = len(re.findall(r"[가-힣A-Za-z0-9\s]", cleaned))
    meaningful_ratio = meaningful / total

    # 길이 점수 (100자 이상이면 1.0)
    length_score = min(total / 100, 1.0)

    score = (
        korean_ratio * 0.4
        + meaningful_ratio * 0.3
        + length_score * 0.2
        + alnum_ratio * 0.1
    )
    return round(min(max(score, 0.0), 1.0), 4)


def _preprocess_image(image: Image.Image) -> Image.Image:
    """OCR 정확도 향상을 위한 이미지 전처리"""
    # 회색조 변환
    gray = ImageOps.grayscale(image)
    # 대비 향상
    enhanced = ImageEnhance.Contrast(gray).enhance(2.0)
    # 이진화 (threshold)
    binary = enhanced.point(lambda x: 255 if x > 140 else 0)
    # 노이즈 제거
    denoised = binary.filter(ImageFilter.MedianFilter(size=3))
    return denoised


def extract_image_text(file_path: str) -> dict[str, Any]:
    """이미지 파일 OCR (Tesseract kor+eng)"""
    path = Path(file_path)
    if not path.is_file():
        return _empty_result(f"이미지 파일을 찾을 수 없습니다: {file_path}")

    try:
        image = Image.open(path)
        processed = _preprocess_image(image)
        text = pytesseract.image_to_string(processed, lang=OCR_LANG)
        return _success_result(text.strip(), "tesseract_ocr", page_count=1, ocr_used=True)
    except Exception as exc:
        logger.error("이미지 OCR 실패 [%s]: %s", file_path, exc)
        return _empty_result(f"이미지 OCR에 실패했습니다: {exc}")


def _pdf_pages_to_text(pdf: pdfplumber.PDF, start: int, end: int) -> str:
    """PDF 페이지 범위에서 텍스트 추출"""
    parts: list[str] = []
    for i in range(start, min(end, len(pdf.pages))):
        page_text = pdf.pages[i].extract_text()
        if page_text:
            parts.append(page_text.strip())
    return "\n".join(parts)


def _pdf_ocr_pages(file_path: str, page_count: int) -> str:
    """PDF 페이지를 이미지로 변환 후 OCR"""
    parts: list[str] = []
    poppler_kwargs: dict[str, Any] = {}
    if config.POPPLER_PATH:
        poppler_kwargs["poppler_path"] = config.POPPLER_PATH

    for start in range(0, page_count, PDF_PAGE_CHUNK_SIZE):
        end = min(start + PDF_PAGE_CHUNK_SIZE, page_count)
        first_page = start + 1
        last_page = end

        try:
            images = convert_from_path(
                file_path,
                first_page=first_page,
                last_page=last_page,
                **poppler_kwargs,
            )
        except Exception as exc:
            logger.error("PDF 이미지 변환 실패: %s", exc)
            raise RuntimeError(
                f"PDF OCR 변환에 실패했습니다. poppler 설치를 확인하세요: {exc}"
            ) from exc

        for image in images:
            processed = _preprocess_image(image)
            page_text = pytesseract.image_to_string(processed, lang=OCR_LANG)
            if page_text.strip():
                parts.append(page_text.strip())

    return "\n".join(parts)


def extract_pdf_text(file_path: str) -> dict[str, Any]:
    """
    PDF 텍스트 추출 (pdfplumber 우선, 텍스트 없으면 OCR 폴백).
    100페이지 초과 시 50페이지씩 분할 처리.
    """
    path = Path(file_path)
    if not path.is_file():
        return _empty_result(f"PDF 파일을 찾을 수 없습니다: {file_path}")

    try:
        with pdfplumber.open(path) as pdf:
            page_count = len(pdf.pages)
            text_parts: list[str] = []

            if page_count > PDF_PAGE_OCR_LIMIT:
                logger.info("대용량 PDF 분할 처리: %d페이지", page_count)
                for start in range(0, page_count, PDF_PAGE_CHUNK_SIZE):
                    chunk_text = _pdf_pages_to_text(pdf, start, start + PDF_PAGE_CHUNK_SIZE)
                    if chunk_text:
                        text_parts.append(chunk_text)
            else:
                text_parts.append(_pdf_pages_to_text(pdf, 0, page_count))

            full_text = "\n".join(text_parts).strip()

            # 텍스트가 충분하지 않으면 OCR 폴백 (스캔 PDF)
            if len(full_text) < 50:
                logger.info("PDF 텍스트 부족 — OCR 폴백 실행: %s", file_path)
                ocr_text = _pdf_ocr_pages(str(path), page_count)
                if len(ocr_text) > len(full_text):
                    return _success_result(
                        ocr_text,
                        "pdf2image+tesseract",
                        page_count=page_count,
                        ocr_used=True,
                    )

            if full_text:
                return _success_result(
                    full_text,
                    "pdfplumber",
                    page_count=page_count,
                    ocr_used=False,
                )

            # pdfplumber 결과 없음 → OCR
            ocr_text = _pdf_ocr_pages(str(path), page_count)
            return _success_result(
                ocr_text,
                "pdf2image+tesseract",
                page_count=page_count,
                ocr_used=True,
            )

    except Exception as exc:
        logger.error("PDF 텍스트 추출 실패 [%s]: %s", file_path, exc)
        return _empty_result(f"PDF 텍스트 추출에 실패했습니다: {exc}")


def extract_excel_data(file_path: str) -> dict[str, Any]:
    """Excel 시트 데이터 추출 (openpyxl / xlrd)"""
    path = Path(file_path)
    if not path.is_file():
        return _empty_result(f"Excel 파일을 찾을 수 없습니다: {file_path}")

    extension = path.suffix.lower()
    text_parts: list[str] = []
    sheet_count = 0

    try:
        if extension == ".xlsx":
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_count = len(wb.sheetnames)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"[시트: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        text_parts.append("\t".join(cells))
            wb.close()

        elif extension == ".xls":
            import xlrd

            wb = xlrd.open_workbook(str(path))
            sheet_count = wb.nsheets
            for sheet in wb.sheets():
                text_parts.append(f"[시트: {sheet.name}]")
                for row_idx in range(sheet.nrows):
                    cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
                    if any(cells):
                        text_parts.append("\t".join(cells))
        else:
            return _empty_result(f"지원하지 않는 Excel 형식입니다: {extension}")

        full_text = "\n".join(text_parts)
        return _success_result(full_text, "openpyxl" if extension == ".xlsx" else "xlrd", page_count=sheet_count)

    except Exception as exc:
        logger.error("Excel 데이터 추출 실패 [%s]: %s", file_path, exc)
        return _empty_result(f"Excel 데이터 추출에 실패했습니다: {exc}")


def _find_libreoffice() -> str | None:
    """LibreOffice 실행 파일 경로 탐색"""
    if getattr(config, "LIBREOFFICE_PATH", ""):
        if Path(config.LIBREOFFICE_PATH).is_file():
            return config.LIBREOFFICE_PATH
    for candidate in LIBREOFFICE_PATHS:
        if Path(candidate).is_file():
            return candidate
    return shutil.which("soffice")


def _extract_hwp_text_pyhwp(file_path: str) -> str | None:
    """HWP5 직접 텍스트 추출 (pyhwp, LibreOffice 불필요)"""
    try:
        result = subprocess.run(
            [sys.executable, "-m", "hwp5.hwp5txt", file_path],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
            check=False,
        )
        text = (result.stdout or "").strip()
        if text:
            return text
        if result.returncode != 0:
            logger.debug("pyhwp 추출 실패 [%s]: %s", file_path, (result.stderr or "")[:300])
    except Exception as exc:
        logger.debug("pyhwp 사용 불가: %s", exc)
    return None


def _libreoffice_hwp_error_message(stderr: str) -> str:
    """LibreOffice HWP 변환 실패 시 사용자 안내 문구"""
    lower = (stderr or "").lower()
    if "source file could not be loaded" in lower:
        return (
            "LibreOffice가 HWP 파일을 열 수 없습니다. "
            "H2Orestart 확장 설치가 필요하거나, HWP를 PDF로 변환 후 업로드하세요."
        )
    if "platform independent libraries" in lower:
        return (
            "LibreOffice 내부 오류로 HWP 변환에 실패했습니다. "
            "LibreOffice를 재설치하거나 HWP 대신 PDF를 업로드하세요."
        )
    return stderr.strip() or "LibreOffice HWP 변환 실패"


def _extract_hwp_text_libreoffice(file_path: str, soffice: str) -> dict[str, Any]:
    """LibreOffice로 HWP → PDF 변환 후 텍스트 추출"""
    path = Path(file_path)
    last_error = ""

    for infilter in LIBREOFFICE_HWP_FILTERS:
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                cmd = [soffice, "--headless"]
                if infilter:
                    cmd.extend(["--infilter", infilter])
                cmd.extend(
                    [
                        "--convert-to",
                        "pdf:writer_pdf_Export",
                        "--outdir",
                        tmp_dir,
                        str(path),
                    ]
                )
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=120,
                    check=False,
                )

                pdf_files = list(Path(tmp_dir).glob("*.pdf"))
                if pdf_files:
                    pdf_result = extract_pdf_text(str(pdf_files[0]))
                    if pdf_result["success"]:
                        pdf_result["method"] = "libreoffice+pdfplumber"
                        return pdf_result
                    return pdf_result

                last_error = _libreoffice_hwp_error_message(
                    (result.stderr or "") + "\n" + (result.stdout or "")
                )
        except subprocess.TimeoutExpired:
            return _empty_result("HWP 변환 시간이 초과되었습니다.")
        except Exception as exc:
            last_error = str(exc)
            logger.debug("LibreOffice HWP 시도 실패 (filter=%s): %s", infilter, exc)

    return _empty_result(last_error or "HWP → PDF 변환 결과 파일이 없습니다.")


def extract_hwp_text(file_path: str) -> dict[str, Any]:
    """
    HWP 텍스트 추출.
    1) pyhwp(HWP5) 직접 추출 시도
    2) 실패 시 LibreOffice PDF 변환 → pdfplumber
    """
    path = Path(file_path)
    if not path.is_file():
        return _empty_result(f"HWP 파일을 찾을 수 없습니다: {file_path}")

    pyhwp_text = _extract_hwp_text_pyhwp(file_path)
    if pyhwp_text:
        quality = check_ocr_quality(pyhwp_text)
        logger.info("HWP 텍스트 추출 완료 (pyhwp): %s", path.name)
        return {
            "success": True,
            "method": "pyhwp",
            "text": pyhwp_text,
            "page_count": 0,
            "ocr_used": False,
            "quality_score": quality,
            "quality_warn": quality < QUALITY_WARN_THRESHOLD,
        }

    soffice = _find_libreoffice()
    if not soffice:
        msg = (
            "HWP 텍스트를 추출하지 못했습니다. "
            "pyhwp 설치(pip install pyhwp six) 후 SAFE를 재시작하거나, "
            "LibreOffice 설치·HWP를 PDF로 변환해 업로드하세요."
        )
        logger.warning(msg)
        return _empty_result(msg)

    try:
        return _extract_hwp_text_libreoffice(file_path, soffice)
    except Exception as exc:
        logger.error("HWP 텍스트 추출 실패 [%s]: %s", file_path, exc)
        return _empty_result(f"HWP 텍스트 추출에 실패했습니다: {exc}")


def extract_text(file_path: str, data_type: str | None = None) -> dict[str, Any]:
    """
    파일 확장자를 자동 감지하여 적절한 추출 함수 호출.
    ZIP은 data_type이 지원하는 경우 묶음 처리.

    Args:
        file_path: 파일 전체 경로
        data_type: 자료유형 "1"~"4" (ZIP 묶음 시 필요)

    Returns:
        추출 결과 dict (text, method, quality_score 등)
    """
    path = Path(file_path)
    if not path.is_file():
        return _empty_result(f"파일을 찾을 수 없습니다: {file_path}")

    extension = path.suffix.lower()
    logger.info("텍스트 추출 시작: %s (확장자: %s)", path.name, extension)

    if extension == ".zip":
        if not data_type:
            return _empty_result("ZIP 파일은 자료유형을 지정해야 합니다.")
        from bundle import extract_zip_bundle_text

        return extract_zip_bundle_text(str(path), data_type)

    return _extract_single_file(str(path))


def _extract_single_file(file_path: str) -> dict[str, Any]:
    """단일 파일 텍스트 추출 (ZIP 제외)"""
    path = Path(file_path)
    extension = path.suffix.lower()

    extractors = {
        ".pdf": extract_pdf_text,
        ".jpg": extract_image_text,
        ".jpeg": extract_image_text,
        ".png": extract_image_text,
        ".tiff": extract_image_text,
        ".xlsx": extract_excel_data,
        ".xls": extract_excel_data,
        ".hwp": extract_hwp_text,
    }

    extractor = extractors.get(extension)
    if not extractor:
        return _empty_result(f"지원하지 않는 파일 형식입니다: {extension}")

    return extractor(str(path))
