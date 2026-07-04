"""
자료유형별 항목 파싱 모듈
추출된 텍스트에서 핵심 항목을 구조화하여 반환
"""

import hashlib
import logging
import re
from decimal import Decimal
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# 날짜 패턴
_DATE_PATTERN = re.compile(
    r"(\d{4})[.\-/년\s]*(\d{1,2})[.\-/월\s]*(\d{1,2})"
)

# 금액 패턴 (숫자+원, 콤마 포함)
_AMOUNT_PATTERN = re.compile(
    r"([\d,]+(?:\.\d+)?)\s*원|"
    r"([\d,]+(?:\.\d+)?)\s*만\s*원|"
    r"(\d+)\s*억\s*(\d+)?\s*천?\s*만?\s*원?"
)


def _parse_korean_amount(text: str) -> int:
    """
    한국식 금액 표기를 정수(원)로 변환.
    예: "1,500,000원" → 1500000, "150만원" → 1500000, "1억 5천만원" → 150000000
    """
    if not text:
        return 0

    cleaned = text.strip().replace(" ", "")

    # 억/만 복합 표기: "1억5천만원", "1억 5천만원"
    eok_man = re.search(r"(\d+)억\s*(\d+)?천?만?원?", cleaned)
    if eok_man:
        eok = int(eok_man.group(1))
        cheon_man = int(eok_man.group(2)) if eok_man.group(2) else 0
        return eok * 100_000_000 + cheon_man * 10_000_000

    # 만원 표기
    man = re.search(r"([\d,]+)만원", cleaned)
    if man:
        return int(man.group(1).replace(",", "")) * 10_000

    # 억원 단독
    eok_only = re.search(r"(\d+)억원", cleaned)
    if eok_only:
        return int(eok_only.group(1)) * 100_000_000

    # 일반 숫자+원 또는 콤마 숫자
    num = re.search(r"([\d,]+)", cleaned)
    if num:
        return int(num.group(1).replace(",", ""))

    return 0


def _find_amount(text: str, keywords: list[str]) -> int:
    """키워드 인근 금액 추출"""
    for keyword in keywords:
        pattern = re.compile(
            rf"{keyword}\s*[:：]?\s*([\d,]+(?:\.\d+)?\s*원|[\d,]+\s*만\s*원|\d+\s*억[^\\n]*)",
            re.IGNORECASE,
        )
        match = pattern.search(text)
        if match:
            return _parse_korean_amount(match.group(1))

        # 키워드 다음 줄 금액 탐색
        line_pattern = re.compile(rf"{keyword}.*?(\n|$)", re.IGNORECASE)
        line_match = line_pattern.search(text)
        if line_match:
            amount_in_line = re.search(
                r"([\d,]+원|[\d,]+만원|\d+억[^\\n]*)",
                line_match.group(0),
            )
            if amount_in_line:
                return _parse_korean_amount(amount_in_line.group(1))

    return 0


def _find_field(text: str, keywords: list[str]) -> str:
    """키워드 기반 텍스트 필드 추출"""
    for keyword in keywords:
        pattern = re.compile(
            rf"{keyword}\s*[:：]\s*(.+?)(?:\n|$)",
            re.IGNORECASE,
        )
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return ""


def _find_field_flexible(text: str, keywords: list[str]) -> str:
    """띄어쓰기가 삽입된 키워드도 인식 (예: 사 업 명)"""
    for keyword in keywords:
        spaced = r"\s*".join(keyword)
        pattern = re.compile(
            rf"{spaced}\s*[:：]\s*(.+?)(?:\n|$)",
            re.IGNORECASE,
        )
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return _find_field(text, keywords)


def _format_period_date(y: str, m: str, d: str | None, *, end: bool = False) -> str:
    """기간용 날짜 문자열 (YYYY.MM.DD). 일 없으면 시작=01일, 종료=말일."""
    import calendar

    yi, mi = int(y), int(m)
    if d:
        return f"{yi:04d}.{mi:02d}.{int(d):02d}"
    if end:
        last = calendar.monthrange(yi, mi)[1]
        return f"{yi:04d}.{mi:02d}.{last:02d}"
    return f"{yi:04d}.{mi:02d}.01"


def _find_date_range(text: str, keywords: list[str]) -> dict[str, str]:
    """기간(시작~종료) 추출 — 콜론·부터/까지·월만 표기·띄어쓰기 변형 지원."""
    result = {"start": "", "end": ""}
    if not text:
        return result

    # 연.월.일 (일 선택) — OCR 공백 허용
    date_tok = (
        r"(\d{4})\s*[.\-/년]\s*(\d{1,2})\s*"
        r"(?:[.\-/월]\s*(\d{1,2})\s*일?)?"
    )
    sep = r"(?:[~\-–—∼～]|부터)"

    for keyword in keywords:
        spaced = r"\s*".join(keyword)
        key_re = re.compile(rf"(?:{re.escape(keyword)}|{spaced})\s*[:：]?\s*", re.I)
        for km in key_re.finditer(text):
            window = text[km.end() : km.end() + 100]
            match = re.match(
                rf"{date_tok}\s*{sep}\s*{date_tok}\s*(?:까지)?",
                window,
            )
            if not match:
                continue
            y1, m1, d1, y2, m2, d2 = match.groups()
            result["start"] = _format_period_date(y1, m1, d1, end=False)
            result["end"] = _format_period_date(y2, m2, d2, end=True)
            return result

    # 키워드 근처(120자)에서 날짜 쌍 탐색
    for keyword in keywords:
        spaced = r"\s*".join(keyword)
        key_re = re.compile(rf"(?:{re.escape(keyword)}|{spaced})", re.I)
        for km in key_re.finditer(text):
            window = text[km.start() : km.start() + 120]
            dates = re.findall(
                r"(\d{4})\s*[.\-/년]\s*(\d{1,2})\s*(?:[.\-/월]\s*(\d{1,2})\s*일?)?",
                window,
            )
            if len(dates) >= 2:
                y1, m1, d1 = dates[0]
                y2, m2, d2 = dates[1]
                result["start"] = _format_period_date(y1, m1, d1 or None, end=False)
                result["end"] = _format_period_date(y2, m2, d2 or None, end=True)
                return result

    return result


def _extract_year_from_period(period: dict[str, str]) -> str:
    """기간 문자열에서 연도 추출"""
    for key in ("start", "end"):
        val = period.get(key, "")
        match = re.search(r"(\d{4})", val)
        if match:
            return match.group(1)
    return ""


def _extract_settlement_summary(text: str) -> dict[str, int]:
    """정산총괄표에서 총예산·집행·잔액 추출"""
    patterns = [
        re.compile(
            r"(?:계|합\s*계)\s*([\d,]+)\s+([\d,]+)\s+\d+%?\s+([\d,]+)",
            re.IGNORECASE,
        ),
        re.compile(
            r"보조금\s*([\d,]+)\s+([\d,]+)\s+\d+%?\s+([\d,]+)",
            re.IGNORECASE,
        ),
    ]
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            return {
                "total_budget": _parse_korean_amount(match.group(1)),
                "total_executed": _parse_korean_amount(match.group(2)),
                "remaining_amount": _parse_korean_amount(match.group(3)),
            }
    return {}


def _extract_settlement_items(text: str) -> list[dict[str, Any]]:
    """정산 항목별 예산·집행·잔액 추출"""
    items: list[dict[str, Any]] = []
    seen: set[str] = set()

    # 발생사유 표 형식 우선 (강사강의료 등)
    block_match = re.search(
        r"집행잔액\s*발생사유.*?(?=|※\s*발생사유|$)",
        text,
        re.DOTALL | re.IGNORECASE,
    )
    search_text = block_match.group(0) if block_match else text

    table_pattern = re.compile(
        r"(강사강의료|인건비|운영비|인쇄비|홍보비|소모품비)"
        r"\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)",
        re.IGNORECASE,
    )
    for match in table_pattern.finditer(search_text):
        category = match.group(1)
        if category in seen:
            continue
        seen.add(category)
        items.append({
            "category": category,
            "budget": _parse_korean_amount(match.group(2)),
            "executed": _parse_korean_amount(match.group(3)),
            "remaining": _parse_korean_amount(match.group(4)),
        })

    # 요약표 인라인 형식 (강사강의료 3,600,000 ... 600,000)
    inline_pattern = re.compile(
        r"강사강의료\s+([\d,]+)\D+([\d,]+)",
        re.IGNORECASE,
    )
    inline_match = inline_pattern.search(text)
    if inline_match and "강사강의료" not in seen:
        budget = _parse_korean_amount(inline_match.group(1))
        executed = _parse_korean_amount(inline_match.group(2))
        items.append({
            "category": "강사강의료",
            "budget": budget,
            "executed": executed,
            "remaining": max(budget - executed, 0),
        })

    return items


def _extract_remaining_reason(text: str) -> str:
    """집행잔액 발생사유 추출"""
    compact = re.sub(r"\s+", "", text)
    if "비대면교육" in compact and "횟수축소" in compact:
        return "비대면교육으로 인한 횟수 축소"

    reason_match = re.search(
        r"발생사유\s*[:：]?\s*([가-힣A-Za-z0-9\s]+?)(?:\n|※)",
        text,
        re.IGNORECASE,
    )
    if reason_match:
        reason = reason_match.group(1).strip()
        if reason and reason != "(단위:원)" and len(reason) > 3:
            return reason
    return ""


def _extract_execution_items_from_settlement(
    text: str,
    default_year: str = "",
) -> list[dict[str, Any]]:
    """정산보고서 내 집행 세부내역에서 집행일자 추출"""
    items: list[dict[str, Any]] = []
    year = default_year or "2021"

    for match in re.finditer(
        r"(\d{1,2})[.,]\s*(\d{1,2})\s+([\d,]+)\s+(\S+)",
        text,
    ):
        month = int(match.group(1))
        day = int(match.group(2))
        amount = _parse_korean_amount(match.group(3))
        item_name = match.group(4).strip()
        if amount <= 0 or month > 12 or day > 31:
            continue
        date_str = f"{year}-{month:02d}-{day:02d}"
        items.append({
            "date": date_str,
            "raw_date": f"{month}.{day:02d}",
            "item_name": item_name,
            "amount": amount,
            "category": item_name,
        })

    return items


def _calc_ratio(part: int, total: int) -> float:
    """비율 계산 (%)"""
    if total <= 0:
        return 0.0
    return round((part / total) * 100, 2)


_BUDGET_PLAN_SECTION_LABELS = (
    "예산집행계획",
    "예산사용계획",
    "예산 집행 계획",
    "예산 사용 계획",
    "비목별 예산",
    "세목별 예산",
)

_BUDGET_PLAN_ITEM_NAMES = (
    "인건비", "사무관리비", "공공운영비", "운영수당", "임차료", "용역비",
    "행사운영비", "국내여비", "국외여비", "국외업무여비", "업무추진비",
    "시설비", "재료비", "연구용역비", "일반운영비", "여비",
)

_EXEC_DATE_HEADERS = (
    "집행거래일자", "거래일자", "집행일자", "지출일자", "지급일자", "이체일자",
)
# 지방비+자부담 합산 (표준 26열 집행내역서)
_EXEC_AMOUNT_LOCAL_HEADERS = (
    "지방비집행금액", "지방비 집행금액",
)
_EXEC_AMOUNT_SELF_HEADERS = (
    "자부담집행금액", "자부담 집행금액",
)
_EXEC_AMOUNT_HEADERS = (
    "집행금액", "거래금액", "지급금액", "이체금액", "집행액",
)
_EXEC_SEOMOK_HEADERS = (
    "보조세목(통계목)", "보조세목", "통계목",
)
_EXEC_NAME_HEADERS = (
    "적요", "내용", "비목", "세목", "용도", "계정과목", "거래내용", "품목", "거래처명",
)

# 표준 26열 엑셀 (1-based): L=12, P=16, T=20, U=21
_EXEC_COL_DATE = 11       # L
_EXEC_COL_SEOMOK = 15     # P
_EXEC_COL_AMOUNT_LOCAL = 19  # T
_EXEC_COL_AMOUNT_SELF = 20   # U
_EXEC_MIN_COLS = 21


def _split_row_cells(line: str) -> list[str]:
    line = line.strip()
    if not line:
        return []
    if "\t" in line:
        return [c.strip() for c in line.split("\t")]
    if "|" in line:
        return [c.strip() for c in line.split("|")]
    return [c.strip() for c in re.split(r"\s{2,}", line)]


def _parse_date_cell(raw: Any) -> str | None:
    """셀/토큰에서 YYYY-MM-DD 추출 (datetime, YYYYMMDD, 2024.10.1 등)."""
    if raw is None or raw == "":
        return None

    # Excel datetime / date
    if hasattr(raw, "strftime"):
        try:
            return raw.strftime("%Y-%m-%d")
        except (ValueError, OSError):
            pass

    # Excel 날짜 일련번호
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        serial = float(raw)
        if 20000 < serial < 80000:
            try:
                from datetime import datetime, timedelta

                # Excel 1900 date system (openpyxl/windows)
                base = datetime(1899, 12, 30)
                return (base + timedelta(days=serial)).strftime("%Y-%m-%d")
            except (ValueError, OverflowError):
                pass

    s = str(raw).strip()
    # datetime 문자열: 2025-08-25 00:00:00
    m = re.match(
        r"(\d{4})-(\d{1,2})-(\d{1,2})(?:\s+\d{1,2}:\d{2})",
        s,
    )
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(
        r"(\d{4})\s*[.\-/년]\s*(\d{1,2})\s*[.\-/월]\s*(\d{1,2})",
        s,
    )
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def _cell_amount(raw: Any) -> int:
    """엑셀 셀 금액 (숫자·문자)."""
    if raw is None or raw == "":
        return 0
    if isinstance(raw, bool):
        return 0
    if isinstance(raw, (int, float)):
        return int(raw)
    return _parse_korean_amount(str(raw))


def _header_norm(text: str) -> str:
    return re.sub(r"\s+", "", str(text or ""))


def _map_execution_header_cols(headers: list[Any]) -> dict[str, int]:
    """헤더 행에서 일자·세목·지방비·자부담 열 인덱스."""
    cols = {"date": -1, "seomok": -1, "amount_local": -1, "amount_self": -1, "name": -1, "amount": -1}
    for j, h in enumerate(headers):
        hn = _header_norm(h)
        if not hn:
            continue
        if cols["date"] < 0 and any(_header_norm(k) in hn or hn in _header_norm(k) for k in _EXEC_DATE_HEADERS):
            cols["date"] = j
        if cols["seomok"] < 0 and any(_header_norm(k) in hn or hn in _header_norm(k) for k in _EXEC_SEOMOK_HEADERS):
            cols["seomok"] = j
        if cols["amount_local"] < 0 and any(_header_norm(k) in hn for k in _EXEC_AMOUNT_LOCAL_HEADERS):
            cols["amount_local"] = j
        if cols["amount_self"] < 0 and any(_header_norm(k) in hn for k in _EXEC_AMOUNT_SELF_HEADERS):
            cols["amount_self"] = j
        if cols["name"] < 0 and any(_header_norm(k) in hn for k in _EXEC_NAME_HEADERS):
            cols["name"] = j
        if cols["amount"] < 0 and any(_header_norm(k) == hn or hn.endswith(_header_norm(k)) for k in _EXEC_AMOUNT_HEADERS):
            # 지방비/자부담이 아닌 단일 집행금액 열
            if "지방비" not in hn and "자부담" not in hn:
                cols["amount"] = j
    return cols


def _row_execution_item(row: list[Any], cols: dict[str, int]) -> dict[str, Any] | None:
    """매핑된 열로 집행 1건 생성. 금액 = 지방비 + 자부담."""
    date_col = cols.get("date", -1)
    if date_col < 0 or date_col >= len(row):
        return None
    date = _parse_date_cell(row[date_col])
    if not date:
        return None

    amount = 0
    local_col = cols.get("amount_local", -1)
    self_col = cols.get("amount_self", -1)
    if local_col >= 0 or self_col >= 0:
        if 0 <= local_col < len(row):
            amount += _cell_amount(row[local_col])
        if 0 <= self_col < len(row):
            amount += _cell_amount(row[self_col])
    elif 0 <= cols.get("amount", -1) < len(row):
        amount = _cell_amount(row[cols["amount"]])

    if amount <= 0:
        return None

    seomok = ""
    seomok_col = cols.get("seomok", -1)
    if 0 <= seomok_col < len(row) and row[seomok_col] is not None:
        seomok = str(row[seomok_col]).strip()

    name = seomok
    name_col = cols.get("name", -1)
    if not name and 0 <= name_col < len(row) and row[name_col] is not None:
        name = str(row[name_col]).strip()

    return {
        "date": date,
        "item_name": name or "(내역)",
        "amount": amount,
        "seomok": seomok,
        "category": _tag_execution_category(name or seomok),
    }


def _tag_execution_category(name: str) -> str:
    if any(k in name for k in ("인건비", "급여", "인력")):
        return "인건비"
    if any(k in name for k in ("시설", "공사", "건축")):
        return "시설비"
    if any(k in name for k in ("운영", "관리", "임차", "용역")):
        return "운영비"
    return "기타"


def _extract_budget_plan_items(text: str) -> list[dict[str, Any]]:
    """① 예산집행계획·예산사용계획 등 세목(항목)별 계획 금액."""
    if not text:
        return []

    section = text
    for label in _BUDGET_PLAN_SECTION_LABELS:
        idx = text.find(label)
        if idx >= 0:
            section = text[idx : idx + 4000]
            break

    items: list[dict[str, Any]] = []
    seen: set[str] = set()

    # 401-01 시설비 50,000,000
    seomok_pat = re.compile(
        r"(?<!\d)(\d{3}(?:-\d{2})?)\s+([가-힣A-Za-z·()/\s]{2,30}?)\s+([\d,]+)\s*원?"
    )
    for match in seomok_pat.finditer(section):
        name = re.sub(r"\s+", " ", match.group(2)).strip()
        amount = _parse_korean_amount(match.group(3))
        if amount <= 0 or len(name) < 2:
            continue
        key = f"{match.group(1)}|{name}|{amount}"
        if key in seen:
            continue
        seen.add(key)
        items.append({
            "seomok": match.group(1),
            "item_name": name,
            "amount": amount,
        })

    if items:
        return items

    # 인건비 10,000,000
    for name in _BUDGET_PLAN_ITEM_NAMES:
        pat = re.compile(rf"{re.escape(name)}\s*[:：]?\s*([\d,]+)\s*원?")
        for match in pat.finditer(section):
            amount = _parse_korean_amount(match.group(1))
            if amount <= 0:
                continue
            key = f"|{name}|{amount}"
            if key in seen:
                continue
            seen.add(key)
            items.append({"seomok": "", "item_name": name, "amount": amount})

    return items


def _extract_execution_items(text: str) -> list[dict[str, Any]]:
    """
    ② 집행내역 행 추출.
    표준: 집행거래일자 + 보조세목 + 지방비·자부담 집행금액 합.
    """
    if not text:
        return []

    items: list[dict[str, Any]] = []
    seen: set[str] = set()

    def _add_item(item: dict[str, Any] | None) -> None:
        if not item:
            return
        key = f"{item['date']}|{item.get('seomok')}|{item['item_name']}|{item['amount']}"
        if key in seen:
            return
        seen.add(key)
        items.append(item)

    def _add(date: str, name: str, amount: int, seomok: str = "") -> None:
        if amount <= 0 or not date:
            return
        name = re.sub(r"\s+", " ", (name or seomok or "").strip())[:80]
        _add_item({
            "date": date,
            "item_name": name or "(내역)",
            "amount": amount,
            "seomok": seomok,
            "category": _tag_execution_category(name or seomok),
        })

    lines = text.splitlines()

    # 1) 헤더 표 (집행거래일자 / 지방비·자부담 / 보조세목)
    for i, line in enumerate(lines):
        cells = _split_row_cells(line)
        if not cells or "집행거래일자" not in _header_norm("".join(cells)):
            # 헤더가 탭으로만 분리된 경우
            if not any("집행거래일자" in _header_norm(c) for c in cells):
                continue
        cols = _map_execution_header_cols(cells)
        if cols["date"] < 0:
            # 표준 26열 위치 폴백 (L/P/T/U)
            if len(cells) >= _EXEC_MIN_COLS:
                cols = {
                    "date": _EXEC_COL_DATE,
                    "seomok": _EXEC_COL_SEOMOK,
                    "amount_local": _EXEC_COL_AMOUNT_LOCAL,
                    "amount_self": _EXEC_COL_AMOUNT_SELF,
                    "name": -1,
                    "amount": -1,
                }
            else:
                continue

        for row_line in lines[i + 1 :]:
            if row_line.startswith("[시트:"):
                break
            row_cells = _split_row_cells(row_line)
            if len(row_cells) <= cols["date"]:
                # 열 수 부족 시 오른쪽에 빈 칸 패딩
                row_cells = row_cells + [""] * (cols["date"] + 1 - len(row_cells))
            # T/U 열까지 필요할 수 있음
            need = max(
                cols.get("amount_local", 0),
                cols.get("amount_self", 0),
                cols.get("seomok", 0),
                cols["date"],
            )
            if len(row_cells) <= need:
                row_cells = row_cells + [""] * (need + 1 - len(row_cells))
            _add_item(_row_execution_item(row_cells, cols))
        if items:
            return items

    # 2) 날짜(또는 YYYYMMDD) + 적요 + 금액
    for match in re.finditer(
        r"(\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2})\s+(.+?)\s+([\d,]+)\s*원?",
        text,
    ):
        date = _parse_date_cell(match.group(1)) or match.group(1)
        _add(date, match.group(2), _parse_korean_amount(match.group(3)))

    for match in re.finditer(
        r"(?<!\d)(\d{4})(\d{2})(\d{2})(?!\d)\s+(.+?)\s+([\d,]+)\s*원?",
        text,
    ):
        date = f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        _add(date, match.group(4), _parse_korean_amount(match.group(5)))

    if items:
        return items

    # 3) 행 단위: 날짜 토큰 + 금액이 함께 있는 줄
    for line in lines:
        date = None
        for token in _split_row_cells(line) or line.split():
            date = _parse_date_cell(token)
            if date:
                break
        if not date:
            m = re.search(r"(\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}|\d{8})", line)
            if m:
                date = _parse_date_cell(m.group(1))
        if not date:
            continue
        amounts = re.findall(r"([\d,]{3,})\s*원?", line)
        amount = 0
        for a in reversed(amounts):
            amount = _parse_korean_amount(a)
            if amount >= 1000:
                break
        if amount <= 0:
            continue
        name = re.sub(r"\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}|\d{8}|[\d,]+원?", " ", line)
        name = re.sub(r"\s+", " ", name).strip()
        _add(date, name, amount)

    return items


def parse_business_plan(text: str) -> dict[str, Any]:
    """① 사업계획서 파싱"""
    total_budget = _find_amount(text, ["총예산", "총 예산", "사업비", "예산액", "총사업비"])
    labor_cost = _find_amount(text, ["인건비", "인력비"])
    operation_cost = _find_amount(text, ["운영비", "사업비"])
    other_cost = _find_amount(text, ["기타경비", "기타 경비", "기타"])

    budget_plan_items = _extract_budget_plan_items(text)
    plan_items_sum = sum(int(i.get("amount", 0)) for i in budget_plan_items)
    if total_budget == 0:
        total_budget = plan_items_sum or (labor_cost + operation_cost + other_cost)

    return {
        "business_name": _find_field(text, ["사업명", "사업 명칭"]),
        "business_purpose": _find_field(text, ["사업목적", "사업 목적"]),
        "business_period": _find_date_range(text, ["사업기간", "사업 기간", "추진기간"]),
        "total_budget": total_budget,
        "budget_breakdown": {
            "labor_cost": labor_cost,
            "operation_cost": operation_cost,
            "other_cost": other_cost,
        },
        "budget_plan_items": budget_plan_items,
        "labor_ratio": _calc_ratio(labor_cost, total_budget),
        "execution_plan": _find_field(text, ["집행계획", "집행 계획", "집행방안"]),
        "settlement_plan": _find_field(text, ["정산계획", "정산 계획", "정산방안"]),
        "applicant_org": _find_field(text, ["신청기관", "수행기관", "주관기관", "신청 기관"]),
    }


def _execution_result_from_items(
    execution_items: list[dict[str, Any]],
    *,
    business_name: str = "",
    total_budget: int = 0,
) -> dict[str, Any]:
    total_executed = sum(int(i.get("amount", 0)) for i in execution_items)
    remaining = max(total_budget - total_executed, 0)
    out_of_budget: list[dict[str, Any]] = []
    if total_executed > total_budget > 0:
        out_of_budget.append({
            "reason": "총집행액이 총예산을 초과",
            "amount": total_executed - total_budget,
        })
    return {
        "business_name": business_name,
        "total_budget": total_budget,
        "total_executed": total_executed,
        "execution_rate": _calc_ratio(total_executed, total_budget),
        "remaining_budget": remaining,
        "execution_items": execution_items,
        "out_of_budget_items": out_of_budget,
    }


def _load_excel_rows(file_path: str) -> list[list[Any]]:
    """엑셀 전체 행 로드 (read_only 미사용 — dimensions 오류 방지)."""
    path = Path(file_path)
    extension = path.suffix.lower()
    all_rows: list[list[Any]] = []

    if extension in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        try:
            for ws in wb.worksheets:
                max_row = max(ws.max_row or 1, 1)
                max_col = max(ws.max_column or 26, 26)
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_col,
                    values_only=True,
                ):
                    cells = list(row)
                    if any(c is not None and str(c).strip() not in ("", "None") for c in cells):
                        all_rows.append(cells)
        finally:
            wb.close()
    elif extension == ".xls":
        import xlrd

        wb = xlrd.open_workbook(str(path))
        for sheet in wb.sheets():
            for row_idx in range(sheet.nrows):
                cells = [
                    sheet.cell_value(row_idx, col)
                    for col in range(min(max(sheet.ncols, 26), 26))
                ]
                if any(str(c).strip() not in ("", "None") for c in cells):
                    all_rows.append(cells)
    else:
        raise ValueError(f"지원하지 않는 엑셀 형식: {extension}")

    return all_rows


def _standard_exec_cols() -> dict[str, int]:
    return {
        "date": _EXEC_COL_DATE,
        "seomok": _EXEC_COL_SEOMOK,
        "amount_local": _EXEC_COL_AMOUNT_LOCAL,
        "amount_self": _EXEC_COL_AMOUNT_SELF,
        "name": -1,
        "amount": -1,
    }


def _extract_items_from_excel_rows(all_rows: list[list[Any]]) -> tuple[list[dict[str, Any]], str]:
    """행 목록에서 집행 건 추출. 헤더 매핑 + L/P/T/U 고정열 병행."""
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    business_name = ""

    def _push(item: dict[str, Any] | None) -> None:
        if not item:
            return
        key = f"{item['date']}|{item.get('seomok')}|{item['amount']}"
        if key in seen:
            return
        seen.add(key)
        items.append(item)

    # 1) 헤더 행 기준
    header_idx = -1
    cols = _standard_exec_cols()
    for i, row in enumerate(all_rows):
        headers = [str(c) if c is not None else "" for c in row]
        if not any("집행거래일자" in _header_norm(h) for h in headers):
            continue
        header_idx = i
        mapped = _map_execution_header_cols(headers)
        cols = _standard_exec_cols()
        if mapped["date"] >= 0:
            cols["date"] = mapped["date"]
        if mapped["seomok"] >= 0:
            cols["seomok"] = mapped["seomok"]
        if mapped["amount_local"] >= 0:
            cols["amount_local"] = mapped["amount_local"]
        if mapped["amount_self"] >= 0:
            cols["amount_self"] = mapped["amount_self"]
        if mapped["name"] >= 0:
            cols["name"] = mapped["name"]
        break

    data_rows = all_rows[header_idx + 1 :] if header_idx >= 0 else all_rows

    for data_row in data_rows:
        padded = list(data_row) + [None] * max(0, 26 - len(data_row))
        # 헤더 반복 행 스킵
        if any("집행거래일자" in _header_norm(str(c or "")) for c in padded):
            continue
        item = _row_execution_item(padded, cols)
        _push(item)
        if item and not business_name and len(padded) > 3 and padded[3]:
            business_name = str(padded[3]).strip()

    # 2) 고정열(L/P/T/U)로 한 번 더 — 헤더 매핑이 빗나간 경우 보완
    fixed = _standard_exec_cols()
    for data_row in all_rows:
        padded = list(data_row) + [None] * max(0, 26 - len(data_row))
        if any("집행거래일자" in _header_norm(str(c or "")) for c in padded):
            continue
        # L열이 날짜일 때만
        if _parse_date_cell(padded[_EXEC_COL_DATE] if len(padded) > _EXEC_COL_DATE else None):
            _push(_row_execution_item(padded, fixed))

    return items, business_name


def parse_execution_detail_excel(file_path: str) -> dict[str, Any]:
    """
    ② 집행내역서 엑셀 전용 파싱.
    L열 집행거래일자, P열 보조세목(통계목), T+U열 지방비·자부담 집행금액.
    """
    path = Path(file_path)
    result = _execution_result_from_items([])
    result["_source_file"] = str(path)

    if not path.is_file():
        logger.warning("집행내역 엑셀 없음: %s", path)
        return result

    try:
        all_rows = _load_excel_rows(str(path))
    except Exception as exc:
        logger.warning("집행내역 엑셀 로드 실패 (%s): %s", path.name, exc)
        return result

    items, business_name = _extract_items_from_excel_rows(all_rows)
    logger.info(
        "집행내역 엑셀 파싱 — 행=%d, 집행건=%d, 파일=%s",
        len(all_rows),
        len(items),
        path.name,
    )
    if not items:
        logger.warning(
            "집행내역 엑셀에서 집행거래일자를 찾지 못함 — 파일=%s, 샘플행=%s",
            path.name,
            all_rows[:2] if all_rows else [],
        )

    out = _execution_result_from_items(items, business_name=business_name)
    out["_source_file"] = str(path)
    return out


def parse_execution_detail(text: str) -> dict[str, Any]:
    """② 집행내역서 파싱 (텍스트)"""
    total_budget = _find_amount(text, ["총예산", "총 예산", "예산액", "배정예산"])
    total_executed = _find_amount(text, ["총집행", "총 집행", "집행합계", "집행액", "집행 합계"])

    execution_items = _extract_execution_items(text)
    item_sum = sum(int(i.get("amount", 0)) for i in execution_items)
    if total_executed == 0:
        total_executed = item_sum

    result = _execution_result_from_items(
        execution_items,
        business_name=_find_field(text, ["사업명", "사업 명칭"]),
        total_budget=total_budget,
    )
    if total_executed > result["total_executed"]:
        result["total_executed"] = total_executed
        result["execution_rate"] = _calc_ratio(total_executed, total_budget)
        result["remaining_budget"] = max(total_budget - total_executed, 0)
    return result


def _merge_execution_results(results: list[dict[str, Any]]) -> dict[str, Any]:
    """ZIP 등 다중 파일 파싱 결과 병합"""
    valid = [r for r in results if r]
    if not valid:
        return parse_execution_detail("")
    if len(valid) == 1:
        return valid[0]

    merged: dict[str, Any] = {
        "business_name": "",
        "total_budget": 0,
        "total_executed": 0,
        "execution_rate": 0.0,
        "remaining_budget": 0,
        "execution_items": [],
        "out_of_budget_items": [],
    }

    for r in valid:
        if not merged["business_name"] and r.get("business_name"):
            merged["business_name"] = r["business_name"]
        merged["total_budget"] = max(merged["total_budget"], r.get("total_budget", 0))
        merged["total_executed"] = max(merged["total_executed"], r.get("total_executed", 0))
        merged["execution_items"].extend(r.get("execution_items", []))
        merged["out_of_budget_items"].extend(r.get("out_of_budget_items", []))

    item_sum = sum(int(i.get("amount", 0)) for i in merged["execution_items"])
    if item_sum > merged["total_executed"]:
        merged["total_executed"] = item_sum

    merged["remaining_budget"] = max(merged["total_budget"] - merged["total_executed"], 0)
    merged["execution_rate"] = _calc_ratio(merged["total_executed"], merged["total_budget"])
    return merged


def parse_execution_detail_bundle(text: str) -> dict[str, Any]:
    """ZIP 묶음 집행내역 — 파일 구분자 기준 분할 후 병합 파싱"""
    chunks = [
        c.strip()
        for c in re.split(r"^===== .+? =====\s*\n", text, flags=re.MULTILINE)
        if c.strip()
    ]
    if len(chunks) <= 1:
        return parse_execution_detail(text)
    return _merge_execution_results([parse_execution_detail(c) for c in chunks])


def _compute_file_hash(file_path: str) -> str:
    """파일 SHA256 해시 계산"""
    path = Path(file_path)
    if not path.is_file():
        return ""
    sha256 = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def parse_expenditure_proof(text: str, file_path: str = "") -> dict[str, Any]:
    """③ 지출증빙자료 파싱"""
    proof_list: list[dict[str, Any]] = []

    # 증빙 항목 패턴
    proof_patterns = [
        re.compile(
            r"(?:증빙|영수|세금계산서|계산서)\s*(?:번호|No)?\s*[:：]?\s*(\S+).*?"
            r"(?:금액|합계)\s*[:：]?\s*([\d,]+원|[\d,]+만원)",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            r"(\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2})\s+"
            r"(.+?)\s+"
            r"([\d,]+원|[\d,]+만원)",
        ),
    ]

    proof_no = 1
    for pattern in proof_patterns:
        for match in pattern.finditer(text):
            if len(match.groups()) >= 2:
                if len(match.groups()) == 2:
                    proof_list.append({
                        "proof_no": match.group(1),
                        "issue_date": "",
                        "amount": _parse_korean_amount(match.group(2)),
                        "supplier": "",
                        "item_name": "",
                        "file_hash": _compute_file_hash(file_path) if file_path else "",
                    })
                else:
                    proof_list.append({
                        "proof_no": str(proof_no),
                        "issue_date": match.group(1),
                        "amount": _parse_korean_amount(match.group(3)),
                        "supplier": "",
                        "item_name": match.group(2).strip(),
                        "file_hash": _compute_file_hash(file_path) if file_path else "",
                    })
                    proof_no += 1

    # 공급자명 추출
    supplier = _find_field(text, ["공급자", "상호", "거래처", "공급자명"])
    for proof in proof_list:
        if not proof["supplier"]:
            proof["supplier"] = supplier

    total_amount = sum(p["amount"] for p in proof_list)

    # 유효하지 않은 증빙 (금액 0 또는 필수정보 누락)
    invalid_proofs = [
        p for p in proof_list
        if p["amount"] <= 0 or not p.get("issue_date") and not p.get("proof_no")
    ]

    return {
        "proof_list": proof_list,
        "total_amount": total_amount,
        "duplicate_detected": [],
        "invalid_proofs": invalid_proofs,
    }


def _merge_expenditure_proof_results(results: list[dict[str, Any]]) -> dict[str, Any]:
    """ZIP 등 다중 파일 지출증빙 파싱 결과 병합"""
    valid = [r for r in results if r]
    if not valid:
        return parse_expenditure_proof("")
    if len(valid) == 1:
        return valid[0]

    merged: dict[str, Any] = {
        "proof_list": [],
        "total_amount": 0,
        "duplicate_detected": [],
        "invalid_proofs": [],
    }
    for r in valid:
        merged["proof_list"].extend(r.get("proof_list", []))
        merged["invalid_proofs"].extend(r.get("invalid_proofs", []))
    merged["total_amount"] = sum(int(p.get("amount", 0)) for p in merged["proof_list"])
    return merged


def parse_expenditure_proof_bundle(text: str, file_path: str = "") -> dict[str, Any]:
    """ZIP 묶음 지출증빙 — 파일 구분자 기준 분할 후 병합 파싱"""
    chunks = [
        c.strip()
        for c in re.split(r"^===== .+? =====\s*\n", text, flags=re.MULTILINE)
        if c.strip()
    ]
    if len(chunks) <= 1:
        return parse_expenditure_proof(text, file_path)
    return _merge_expenditure_proof_results(
        [parse_expenditure_proof(c, file_path) for c in chunks]
    )


def parse_settlement_report(text: str) -> dict[str, Any]:
    """④ 정산보고서 파싱"""
    business_period = _find_date_range(text, ["사업기간", "사업 기간", "추진기간"])
    settlement_period = _find_date_range(text, ["정산기간", "정산 기간"])
    if not business_period.get("start"):
        business_period = settlement_period

    summary = _extract_settlement_summary(text)
    total_budget = summary.get("total_budget") or _find_amount(
        text, ["총예산", "총 예산", "예산액", "배정예산", "예 산 액"],
    )
    total_executed = summary.get("total_executed") or _find_amount(
        text, ["총집행", "총 집행", "집행합계", "집행액", "지 출 액"],
    )
    remaining = summary.get("remaining_amount") or max(total_budget - total_executed, 0)
    refund_amount = _find_amount(text, ["반납금", "반납 금액", "환수금", "반환금"])
    if refund_amount == 0 and remaining > 0:
        refund_amount = remaining

    settlement_items = _extract_settlement_items(text)
    if not settlement_items:
        category_keywords = ["인건비", "운영비", "기타경비", "기타", "사업비", "강사강의료"]
        for category in category_keywords:
            budget = _find_amount(text, [f"{category} 예산", f"{category} 배정", category])
            executed = _find_amount(text, [f"{category} 집행", f"{category} 사용", f"{category} 지출"])
            if budget > 0 or executed > 0:
                settlement_items.append({
                    "category": category,
                    "budget": budget,
                    "executed": executed,
                    "remaining": max(budget - executed, 0),
                })

    default_year = _extract_year_from_period(business_period)
    execution_items = _extract_execution_items_from_settlement(text, default_year)

    refund_plan = _find_field_flexible(text, ["반납계획", "반납 계획", "반환계획"])
    settlement_deadline = _find_field_flexible(text, ["정산기한", "정산 기한", "정산마감일"])
    remaining_reason = _extract_remaining_reason(text)

    return {
        "business_name": _find_field_flexible(text, ["사업명", "사업 명칭"]),
        "business_period": business_period,
        "settlement_period": settlement_period,
        "total_budget": total_budget,
        "total_executed": total_executed,
        "remaining_amount": remaining,
        "execution_rate": _calc_ratio(total_executed, total_budget),
        "refund_amount": refund_amount,
        "refund_plan": refund_plan,
        "settlement_deadline": settlement_deadline,
        "remaining_reason": remaining_reason,
        "settlement_items": settlement_items,
        "execution_items": execution_items,
    }


def detect_duplicate_proof(files: list[str]) -> list[dict[str, Any]]:
    """
    중복 증빙 탐지 (SHA256 해시 비교).

    Args:
        files: 파일 경로 목록

    Returns:
        중복 탐지 결과 목록
    """
    hash_map: dict[str, list[str]] = {}
    duplicates: list[dict[str, Any]] = []

    for file_path in files:
        path = Path(file_path)
        if not path.is_file():
            logger.warning("파일을 찾을 수 없음 (중복 탐지 건너뜀): %s", file_path)
            continue

        file_hash = _compute_file_hash(str(path))
        if not file_hash:
            continue

        if file_hash in hash_map:
            hash_map[file_hash].append(str(path))
        else:
            hash_map[file_hash] = [str(path)]

    for file_hash, paths in hash_map.items():
        if len(paths) > 1:
            duplicates.append({
                "file_hash": file_hash,
                "files": paths,
                "count": len(paths),
            })
            logger.warning("중복 증빙 탐지: %s", paths)

    return duplicates


def parse_document(
    text: str,
    data_type: str,
    file_path: str = "",
) -> dict[str, Any]:
    """
    자료유형에 맞는 파서 자동 호출.

    Args:
        text: 추출된 텍스트
        data_type: "1"~"4"
        file_path: 원본 파일 경로 (증빙 해시용, 선택)

    Returns:
        파싱 결과 dict
    """
    if not text or not text.strip():
        # 엑셀 집행내역은 파일 직접 파싱 가능
        if not (data_type == "2" and file_path):
            raise ValueError("파싱할 텍스트가 비어 있습니다.")

    logger.info("문서 파싱 시작 — 자료유형: %s", data_type)

    if data_type == "2":
        path = Path(file_path) if file_path else None
        if path and path.is_file() and path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
            result = parse_execution_detail_excel(str(path))
            result["_source_file"] = str(path)
            n_items = len(result.get("execution_items") or [])
            logger.info("문서 파싱 완료 — 자료유형: 2 (엑셀 %d건) %s", n_items, path.name)
            if n_items > 0:
                return result
            logger.warning("엑셀 직접 파싱 0건 — 텍스트 파서 재시도: %s", path.name)
        result = parse_execution_detail(text or "")
        if path:
            result["_source_file"] = str(path)
        logger.info(
            "문서 파싱 완료 — 자료유형: 2 (텍스트 %d건)",
            len(result.get("execution_items") or []),
        )
        return result

    parsers = {
        "1": parse_business_plan,
        "3": lambda t: parse_expenditure_proof_bundle(t, file_path)
        if "===== " in t and " =====" in t
        else parse_expenditure_proof(t, file_path),
        "4": parse_settlement_report,
    }
    parser = parsers.get(data_type)
    if not parser:
        raise ValueError(f"지원하지 않는 자료유형입니다: {data_type}")

    result = parser(text)
    logger.info("문서 파싱 완료 — 자료유형: %s", data_type)
    return result
