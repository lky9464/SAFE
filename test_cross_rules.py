"""교차 항목 규칙 (JC-01, X07) 단위 테스트"""

from __future__ import annotations

import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(_ROOT))

from cross_rules import apply_cross_item_rule, rule_jc01, rule_x07  # noqa: E402

PASS = "[PASS]"
FAIL = "[FAIL]"


def _ok(cond: bool, label: str) -> bool:
    print(f"  {PASS if cond else FAIL} {label}")
    return cond


def test_x07_pass() -> bool:
    print("\n=== X07: 기간 내 집행 ===")
    parsed = {
        "business_period": {"start": "2025.01.01", "end": "2025.12.31"},
        "execution_items": [
            {"date": "2025-03-15", "item_name": "시설비", "amount": 1_000_000},
            {"date": "2025.06.01", "item_name": "자재", "amount": 500_000},
        ],
    }
    r = rule_x07(parsed)
    return _ok(r is not None and r["judge_result"] == "P", f"P: {r['judge_result']}")


def test_x07_fail_outside() -> bool:
    print("\n=== X07: 기간 외 집행 ===")
    parsed = {
        "business_period": {"start": "2025-01-01", "end": "2025-06-30"},
        "execution_items": [
            {"date": "2025-03-01", "item_name": "정상", "amount": 100},
            {"date": "2025-08-01", "item_name": "기간외", "amount": 200},
        ],
    }
    r = rule_x07(parsed)
    return _ok(r is not None and r["judge_result"] == "F", f"F: {(r or {}).get('judge_reason', '')[:40]}")


def test_x07_warn_no_period() -> bool:
    print("\n=== X07: 사업기간 없음 ===")
    r = rule_x07({"execution_items": [{"date": "2025-01-01", "amount": 1}]})
    return _ok(r is not None and r["judge_result"] == "W", f"W: {r and r['judge_reason'][:40]}")


def test_jc01_pass() -> bool:
    print("\n=== JC-01: 계획 ≥ 집행 ===")
    parsed = {
        "plan_total_budget": 10_000_000,
        "total_executed": 8_000_000,
        "budget_breakdown": {"labor_cost": 0},
    }
    r = rule_jc01(parsed)
    return _ok(
        r is not None and r["judge_result"] == "P" and "8,000,000" in r["extracted_val"],
        f"P: {r and r['extracted_val']}",
    )


def test_jc01_fail_over() -> bool:
    print("\n=== JC-01: 집행 초과 ===")
    parsed = {
        "plan_total_budget": 5_000_000,
        "total_executed": 6_000_000,
    }
    r = rule_jc01(parsed)
    return _ok(r is not None and r["judge_result"] == "F", f"F: {(r or {}).get('judge_reason', '')[:40]}")


def test_period_formats() -> bool:
    print("\n=== 사업기간 파싱 형식 ===")
    from parser import _find_date_range

    samples = [
        "사업기간: 2025.01.01 ~ 2025.12.31",
        "사업기간 2025년 1월 1일부터 2025년 12월 31일까지",
        "사 업 기 간 : 2025. 3. 1. ~ 2025. 11. 30.",
        "사업기간: 2025.01 ~ 2025.12",
    ]
    ok = True
    for s in samples:
        found = _find_date_range(s, ["사업기간", "사업 기간"])
        good = bool(found.get("start") and found.get("end"))
        ok = ok and good
        print(f"  {'[PASS]' if good else '[FAIL]'} {s[:40]} -> {found}")
    return ok


def test_x07_from_combined_text() -> bool:
    print("\n=== X07: combined_text 에서 기간 복구 ===")
    parsed = {
        "business_period": {"start": "", "end": ""},
        "combined_text": (
            "=== 사업계획서 ===\n"
            "사업기간: 2025.01.01 ~ 2025.12.31\n"
            "=== 집행내역서 ===\n"
        ),
        "execution_items": [
            {"date": "2025-06-01", "item_name": "공사비", "amount": 1000},
        ],
    }
    r = rule_x07(parsed)
    return _ok(r is not None and r["judge_result"] == "P", f"P from combined: {r and r['extracted_val']}")


def test_execution_trade_date_header() -> bool:
    print("\n=== 집행거래일자 헤더 파싱 ===")
    from parser import _extract_execution_items

    text = (
        "집행거래일자\t적요\t집행금액\n"
        "2024-11-05\t시설공사비\t1,200,000\n"
        "20241020\t자재비\t300000\n"
    )
    items = _extract_execution_items(text)
    ok = len(items) >= 2 and items[0]["date"].startswith("2024")
    print(f"  {'[PASS]' if ok else '[FAIL]'} items={items}")
    return ok


def test_standard_26col_excel_layout() -> bool:
    print("\n=== 표준 26열 (L/P/T+U) ===")
    from parser import _extract_execution_items, parse_execution_detail_excel

    headers = [""] * 26
    headers[0] = "순번"
    headers[11] = "집행거래일자"
    headers[15] = "보조세목(통계목)"
    headers[19] = "지방비 집행금액"
    headers[20] = "자부담 집행금액"
    row1 = [""] * 26
    row1[0] = "1"
    row1[11] = "2025-08-25"
    row1[15] = "시설비"
    row1[19] = "450,000,000"
    row1[20] = "0"
    row2 = [""] * 26
    row2[0] = "2"
    row2[11] = "2025-09-12 00:00:00"
    row2[15] = "시설비"
    row2[19] = "87000000"
    row2[20] = "200000000"

    text = "\t".join(headers) + "\n" + "\t".join(row1) + "\n" + "\t".join(row2)
    items = _extract_execution_items(text)
    ok_text = (
        len(items) == 2
        and items[0]["date"] == "2025-08-25"
        and items[0]["amount"] == 450_000_000
        and items[0]["seomok"] == "시설비"
        and items[1]["amount"] == 287_000_000
    )
    print(f"  {'[PASS]' if ok_text else '[FAIL]'} text items={items}")

    # 실제 xlsx
    try:
        from openpyxl import Workbook
        from pathlib import Path
        import tempfile

        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append([
            1, "", "", "자재창고", "", "", "", "", "", "", "",
            __import__("datetime").date(2025, 8, 25),
            "", "", "", "시설비", "", "", "",
            450_000_000, 0,
        ])
        # ensure 26 cols - append pads differently; write by cell
        ws.cell(2, 12, __import__("datetime").date(2025, 8, 25))  # L
        ws.cell(2, 16, "시설비")  # P
        ws.cell(2, 20, 450_000_000)  # T
        ws.cell(2, 21, 0)  # U
        ws.cell(3, 12, __import__("datetime").datetime(2025, 9, 12))
        ws.cell(3, 16, "시설비")
        ws.cell(3, 20, 87_000_000)
        ws.cell(3, 21, 200_000_000)

        tmp = Path(tempfile.gettempdir()) / "safe_exec_test.xlsx"
        wb.save(tmp)
        result = parse_execution_detail_excel(str(tmp))
        items_x = result.get("execution_items") or []
        ok_xlsx = (
            len(items_x) >= 2
            and items_x[0]["date"] == "2025-08-25"
            and items_x[0]["amount"] == 450_000_000
            and items_x[1]["amount"] == 287_000_000
        )
        print(f"  {'[PASS]' if ok_xlsx else '[FAIL]'} xlsx items={items_x}")
        tmp.unlink(missing_ok=True)
        return ok_text and ok_xlsx
    except Exception as exc:
        print(f"  [FAIL] xlsx: {exc}")
        return False


def test_budget_plan_items() -> bool:
    print("\n=== 예산집행계획 세목 파싱 ===")
    from parser import _extract_budget_plan_items, parse_business_plan

    text = (
        "예산집행계획\n"
        "401-01 시설비 50,000,000\n"
        "201-01 사무관리비 2,000,000\n"
    )
    items = _extract_budget_plan_items(text)
    plan = parse_business_plan(text)
    ok = len(items) >= 2 and plan["total_budget"] >= 52_000_000
    print(f"  {'[PASS]' if ok else '[FAIL]'} items={items} total={plan['total_budget']}")
    return ok


def test_jc01_with_plan_items() -> bool:
    print("\n=== JC-01: 예산집행계획 항목 대조 ===")
    parsed = {
        "budget_plan_items": [
            {"seomok": "401-01", "item_name": "시설비", "amount": 50_000_000},
            {"seomok": "201-01", "item_name": "사무관리비", "amount": 2_000_000},
        ],
        "execution_items": [
            {"date": "2024-11-05", "item_name": "시설공사", "amount": 48_000_000},
        ],
        "total_executed": 48_000_000,
    }
    r = rule_jc01(parsed)
    ok = r is not None and r["judge_result"] == "P" and "2개 항목" in r["extracted_val"]
    print(f"  {'[PASS]' if ok else '[FAIL]'} {r}")
    return ok


def test_x07_trade_dates() -> bool:
    print("\n=== X07: 집행거래일자 대조 ===")
    parsed = {
        "business_period": {"start": "2024.10.01", "end": "2025.10.31"},
        "combined_text": (
            "=== 집행내역서 ===\n"
            "집행거래일자\t적요\t집행금액\n"
            "2024-11-05\t공사비\t1000000\n"
            "2025-01-10\t자재\t500000\n"
        ),
        "execution_items": [],
    }
    r = rule_x07(parsed)
    ok = r is not None and r["judge_result"] == "P"
    print(f"  {'[PASS]' if ok else '[FAIL]'} {r}")
    return ok


def test_jc01_via_dispatcher() -> bool:
    print("\n=== apply_cross_item_rule ===")
    r = apply_cross_item_rule(
        "JC-01",
        {"plan_total_budget": 100, "total_executed": 50},
    )
    r2 = apply_cross_item_rule("JC-99", {"plan_total_budget": 100})
    return _ok(r is not None and r2 is None, "JC-01 only")


def main() -> int:
    print("# 교차 항목 규칙 테스트")
    results = [
        test_x07_pass(),
        test_x07_fail_outside(),
        test_x07_warn_no_period(),
        test_jc01_pass(),
        test_jc01_fail_over(),
        test_period_formats(),
        test_x07_from_combined_text(),
        test_execution_trade_date_header(),
        test_standard_26col_excel_layout(),
        test_budget_plan_items(),
        test_jc01_with_plan_items(),
        test_x07_trade_dates(),
        test_jc01_via_dispatcher(),
    ]
    passed = sum(results)
    print(f"\n결과: {passed}/{len(results)}")
    return 0 if passed == len(results) else 1


if __name__ == "__main__":
    sys.exit(main())
