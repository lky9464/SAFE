"""Phase 1 Excel 산출물 생성 (일제점검 체크리스트 기반)"""
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from phase1_inspection_items import (
    ALL_ITEMS,
    CROSS_RULES,
    DOC_MAPPING,
    GOLDEN_ROWS,
    JUDGMENT_MAPPING,
    SCENARIO_SUMMARY,
)

OUT_DIR = Path(__file__).resolve().parent.parent / "docs" / "phase1"

LEGACY_ENGLISH_ALIASES = (
    "02_golden_cases.xlsx",
    "05_mapping_sample30.xlsx",
    "06_profile_item_matrix.xlsx",
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def autosize(ws, max_width: int = 48) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row_idx, col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _save_wb(wb: Workbook, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(path.stem + "_new.xlsx")
        wb.save(alt)
        print(f"WARN: locked -> {alt.name}")
        return alt


def remove_legacy_aliases() -> list[Path]:
    removed: list[Path] = []
    for name in LEGACY_ENGLISH_ALIASES:
        path = OUT_DIR / name
        if path.is_file():
            path.unlink()
            removed.append(path)
    return removed


def build_inspection_checklist() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "일제점검항목"
    headers = [
        "항목ID", "편성목", "세목", "점검사항", "점검서류",
        "required_docs", "통합분류", "N/A조건",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in ALL_ITEMS:
        ws.append(list(row))
        for c in range(1, len(headers) + 1):
            ws.cell(ws.max_row, c).alignment = WRAP

    ws2 = wb.create_sheet("구비서류매핑")
    ws2.append(["구비서류", "SAFE", "포함내용", "비고"])
    style_header(ws2, 1, 4)
    for r in DOC_MAPPING:
        ws2.append(list(r))

    ws3 = wb.create_sheet("판정매핑")
    ws3.append(["실무(일제점검)", "SAFE자동", "최종보고(3값)", "비고"])
    style_header(ws3, 1, 4)
    for r in JUDGMENT_MAPPING:
        ws3.append(list(r))

    for sheet in (ws, ws2, ws3):
        autosize(sheet)
    return _save_wb(wb, OUT_DIR / "08_일제점검_체크리스트_항목.xlsx")


def build_golden() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "골든케이스"
    headers = [
        "시나리오ID", "사업명(가명)", "②집행세목", "①②③④",
        "항목ID", "점검항목", "SAFE기대", "최종보고(3값)", "사유", "팀확인",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for r in GOLDEN_ROWS:
        ws.append(list(r))
        for c in range(1, len(headers) + 1):
            ws.cell(ws.max_row, c).alignment = WRAP

    ws2 = wb.create_sheet("시나리오요약")
    ws2.append(["ID", "설명", "101집행", "주요세목", "③", "검증포인트"])
    style_header(ws2, 1, 6)
    for r in SCENARIO_SUMMARY:
        ws2.append(list(r))

    ws3 = wb.create_sheet("구V1-V3참고")
    ws3.append(["구ID", "신ID", "비고"])
    style_header(ws3, 1, 3)
    for r in [
        ("V1", "G1", "공사만·101 없음"),
        ("V2", "G2", "인건비+③완비"),
        ("V3", "G6", "③ 없음"),
    ]:
        ws3.append(list(r))

    for sheet in (ws, ws2, ws3):
        autosize(sheet)
    return _save_wb(wb, OUT_DIR / "02_골든케이스.xlsx")


def build_mapping() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "일제점검47항목"
    headers = [
        "항목ID", "편성목", "세목", "점검사항", "required_docs",
        "점검서류", "통합분류", "N/A조건",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in ALL_ITEMS:
        ws.append([row[0], row[1], row[2], row[3], row[5], row[4], row[6], row[7]])
        for c in range(1, len(headers) + 1):
            ws.cell(ws.max_row, c).alignment = WRAP
    autosize(ws)
    return _save_wb(wb, OUT_DIR / "05_일제점검_통합매핑표.xlsx")


def build_matrix() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "세목적용매트릭스"
    seomok_cols = [
        "101", "201-01", "201-32", "201-33", "201-34", "201-02",
        "202-01", "202-03", "203-01~04", "401-01",
    ]
    headers = ["항목ID", "점검항목", "required_docs"] + seomok_cols + ["③없음", "비고"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    def seomok_for_item(item_id: str) -> dict[str, bool]:
        mapping = {
            "J101": "101", "J20101": "201-01", "J20132": "201-32",
            "J20133": "201-33",
            "J20134": "201-34", "J20102": "201-02", "J20201": "202-01",
            "J20203": "202-03", "J203": "203-01~04", "J40101": "401-01",
        }
        for prefix, col in mapping.items():
            if item_id.startswith(prefix):
                return {col: True}
        return {}

    key_ids = [
        "J101-02", "J101-01", "J20132-02", "J20133-03", "J40101-02",
        "J20101-02", "JC-01", "JC-02", "X07",
    ]
    notes = {
        "J101-02": "② 101 없으면 N/A",
        "JC-01": "①② 규칙 적용",
        "J20101-02": "③ 없으면 N/A",
        "JC-02": "② 비목 적정",
        "X07": "① 기간+② 집행일, ① 없으면 W",
    }
    no_proof = {
        "J101-02": "N/A",
        "J20101-02": "N/A",
        "JC-01": "적용",
    }
    id_to_name = {r[0]: r[3] for r in ALL_ITEMS}
    id_to_name.update({r[0]: r[3] for r in CROSS_RULES})
    req_docs = {r[0]: r[5] for r in ALL_ITEMS}
    req_docs.update({r[0]: r[5] for r in CROSS_RULES})

    for item_id in key_ids:
        applies = seomok_for_item(item_id)
        row = [item_id, id_to_name.get(item_id, ""), req_docs.get(item_id, "")]
        for col in seomok_cols:
            row.append("적용" if applies.get(col) else "N/A")
        row.extend([no_proof.get(item_id, ""), notes.get(item_id, "")])
        ws.append(row)
        for c in range(1, len(headers) + 1):
            ws.cell(ws.max_row, c).alignment = WRAP

    ws2 = wb.create_sheet("골든검증")
    ws2.append(["시나리오", "항목ID", "SAFE기대", "검증", "일치"])
    style_header(ws2, 1, 5)
    for sid, _, _, _, iid, _, safe, _, reason, _ in GOLDEN_ROWS:
        if iid:
            ws2.append([sid, iid, safe, reason, ""])
    autosize(ws2)
    autosize(ws)
    return _save_wb(wb, OUT_DIR / "06_프로필_항목_매트릭스.xlsx")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for path in remove_legacy_aliases():
        print(f"removed legacy alias: {path.name}")
    for fn in (build_inspection_checklist, build_golden, build_mapping, build_matrix):
        print(fn())


if __name__ == "__main__":
    main()
