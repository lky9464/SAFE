"""SAFE 개선 로드맵 — Excel(.xlsx) 요약 시트 생성"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT_DIR = Path(__file__).resolve().parent.parent / "docs"
FILES = ["SAFE_개선로드맵.xlsx", "SAFE_improvement_roadmap.xlsx"]


def build() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "로드맵요약"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "SAFE 대규모 개선 로드맵 (요약)"
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal="center")

    ws["A2"] = "작성일: 2026-06-29 | 상태: 기획·요구정리 | 상세: SAFE_개선로드맵.docx 참고"
    ws.merge_cells("A2:E2")

    headers = ["단계", "목표", "주요 작업", "리스크", "비고"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = wrap

    rows = [
        [
            "1단계",
            "요구·도메인·골든케이스 문서화 (코드 전 2~4주)",
            "사업 검토 정의, 통합 체크리스트 분류, 프로필 매트릭스, 판정 유형, 골든 케이스 2~3건",
            "낮음",
            "개발 착수 전 합의 필수",
        ],
        [
            "2단계",
            "사업 프로필 + 항목 적용조건 + N/A",
            "검토 시작 폼, applies_when/exclude_when, 규칙검증 연동, N/A UI",
            "낮음",
            "3단계 전 필수 — 오판 감소",
        ],
        [
            "3단계",
            "사업(case) 단위 4종 통합 검토",
            "case 엔티티, 통합 업로드, 교차 항목, 결과 case_id 관리",
            "중~高",
            "SILO 해소 핵심",
        ],
        [
            "4단계",
            "지식DB 연동 + 판정 고도화",
            "항목-지식 연결, 교차 판정 엔진, 유사도 보조화",
            "高",
            "지식 품질·유지보수",
        ],
    ]
    for ri, row in enumerate(rows, 5):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.alignment = wrap

    ws2 = wb.create_sheet("교차항목예시")
    ws2.append(["점검 내용", "필요 자료"])
    for r in [
        ("계획 예산 vs 집행 합계", "① + ②"),
        ("집행 vs 증빙 금액·건수", "② + ③"),
        ("정산 총괄 vs 집행내역", "② + ④"),
        ("사업기간 vs 집행일·증빙일", "① + ② + ③"),
        ("자부담 비율 (계획 vs 정산)", "① + ④"),
    ]:
        ws2.append(list(r))

    ws3 = wb.create_sheet("1단계체크리스트")
    ws3.append(["완료", "할 일"])
    for item in [
        "실제 검토 사업 2~3건 정리",
        "교차 항목 10개 선정",
        "사업 프로필 10문항 초안",
        "통합 체크리스트 카테고리 합의",
        "DB 마이그레이션 방침",
    ]:
        ws3.append(["", item])

    from openpyxl.utils import get_column_letter

    for ws_ in (ws, ws2, ws3):
        for col_idx in range(1, ws_.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 8
            for row_idx in range(1, ws_.max_row + 1):
                val = ws_.cell(row_idx, col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws_.column_dimensions[col_letter].width = min(max_len + 2, 50)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    first = OUT_DIR / FILES[0]
    wb.save(first)
    for name in FILES[1:]:
        import shutil
        shutil.copy2(first, OUT_DIR / name)
    return first


if __name__ == "__main__":
    print(build())
